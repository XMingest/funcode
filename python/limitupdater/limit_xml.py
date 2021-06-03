# -*- coding: utf-8 -*-
import re
import xml.etree.ElementTree as Et

import openpyxl
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side


class LimitXml:
    @staticmethod
    def parse_range(range_text):
        """
        将形如`(85:160)` `[0.285:0.325]`之类的字符串转换为格式化字典
        >>> LimitXml.parse_range('(85:160)')
        {'cp': ':', 'lmt': ['85', '160'], 'lp': '(', 'rp': ')'}
        >>> LimitXml.parse_range('[0.285:0.325]')
        {'cp': ':', 'lmt': ['0.285', '0.325'], 'lp': '[', 'rp': ']'}
        >>> LimitXml.parse_range('\\left( 8 \\to 9 \\right)')
        {'cp': ' \\to ', 'lmt': ['8', '9'], 'lp': '\\\\left( ', 'rp': ' \\right)'}
        >>> LimitXml.parse_range('(8 8, 8 8 8)')
        {'cp': ' ', 'lmt': ['8', '8', '8', '8', '8'], 'lp': '(', 'rp': ')'}
        >>> LimitXml.parse_range('(8, 8 8 8 8)')
        {'cp': ', ', 'lmt': ['8', '8', '8', '8', '8'], 'lp': '(', 'rp': ')'}
        >>> LimitXml.parse_range('(8)')
        {'cp': '', 'lmt': ['8'], 'lp': '(', 'rp': ')'}
        >>> LimitXml.parse_range('...105')
        {'cp': '', 'lmt': ['.105'], 'lp': '..', 'rp': ''}
        >>> LimitXml.parse_range('10...')
        {'cp': '', 'lmt': ['10'], 'lp': '', 'rp': '...'}
        >>> LimitXml.parse_range('10.132.31.105')
        {'cp': '', 'lmt': ['10.132', '.31', '.105'], 'lp': '', 'rp': ''}
        >>> LimitXml.parse_range('.132.31.105')
        {'cp': '', 'lmt': ['.132', '.31', '.105'], 'lp': '', 'rp': ''}
        >>> LimitXml.parse_range('.923-.856')
        {'cp': '', 'lmt': ['.923', '-.856'], 'lp': '', 'rp': ''}

        :param range_text:
        :return:
        """
        result = {}
        nums = [m for m in re.finditer('-?\d*\.?\d+', range_text)]
        result['cp'] = (range_text[nums[0].end(): nums[1].start()]
                        if len(nums) > 1 and nums[0].end() < nums[1].start()
                        else '')
        result['lmt'] = [n.group() for n in nums]
        result['lp'] = range_text[0: nums[0].start()] if nums[0].start() else ''
        result['rp'] = range_text[nums[-1].end():] if nums[-1].end() < len(range_text) else ''
        return result

    INF_LIMIT = 0xFFFF
    MAX_DIGHT_N = 5

    def __init__(self, path):
        """
        存在结构如下的xml
        ```xml
<?xml version="1.0" encoding="GB2312"?>
<MachineTestStandard>
<MachineType>MediaTest 9508B</MachineType>
<ReleaseVersion>1.021</ReleaseVersion>
<ReleaseDate>2020/08/07</ReleaseDate>
<TestItems>
    <TestItem Name="White90Hz" ChromaType="Rvs">
        <AlgorithmItem Name="Chroma" Desc="Chroma">
            <TestStandards Desc="多媒体V6.0" ProductType="Noah" IsActive="1">
                <CriterionItem Name="CCT" ReverseName1="CCT" ReverseName2="CCTTolerance"
                               Desc="色温中心值" Remark="" Type="2" Range="[6000:8000]" />
                <CriterionItem Name="Cx" ReverseName1="C1" ReverseName2="C1Tolerance"
                               Desc="色坐标x" Remark="" Type="2" Range="[0.285:0.325]" />
                <CriterionItem Name="Cy" ReverseName1="C2" ReverseName2="C2Tolerance"
                               Desc="色坐标y" Remark="" Type="2" Range="[0.301:0.341]" />
            </TestStandards>
        </AlgorithmItem>
    </TestItem>
</TestItems>
</MachineTestStandard>
        ```
        则从中得到
        ```python
{'White90Hz': {'Chroma': {
                            'CCT': '[6000:8000]',
                            'Cx': '[0.285:0.325]',
                            'Cy': '[0.301:0.341]',
}}}
        ```
        :param path:
        :return:
        """
        self.path = path
        self.origin_root = None
        self.text = None
        self.read_limit_xml()
        # 格式化数据
        self.data = {}
        for screen_item in self.origin_root.findall('.//TestItem[@Name]'):
            screen = screen_item.get('Name')
            for algorithm_item in screen_item.findall('.//AlgorithmItem[@Name]'):
                algorithm = algorithm_item.get('Name')
                for criterion_item in algorithm_item.findall('.//CriterionItem[@Name]'):
                    criterion = criterion_item.get('Name')
                    if 'Range' in criterion_item.keys():
                        limit = criterion_item.get('Range')
                    else:
                        limit = criterion_item.text
                    if limit:
                        self.data.setdefault(screen, {}) \
                            .setdefault(algorithm, {})[criterion] = [limit, None, None,
                                                                     (screen, algorithm, criterion)]
        # 自动修改应用
        self.auto_update()
        return

    def auto_update(self):
        """
        根据一定的规则对门限进行自动合入
        :return:
        """
        for screen, screen_item in self.data.items():
            for algorithm, algorithm_item in screen_item.items():
                for criterion in algorithm_item:
                    # 常见收紧
                    for pattern, size in (
                        ('(?:Center|Inner|Outer)(?:Dark|Light)(?:Line|Region|Spot)Mura(?:Avg|Max)ContrastThr', .002),
                        ('(?:Center|Inner|Outer)(?:Dark|Light)LineMuraAreaThr', 100),
                    ):
                        if re.match(pattern, criterion):
                            self.tighten_range([screen, algorithm, criterion], size, is_auto=True)
                            break
                    else:
                        if re.match('(?:Center|Inner|Outer)(?:Dark|Light)(?:Region|Spot)MuraAreaThr',
                                    criterion):
                            # 面积上限不大于则80不变
                            if float(self.parse_range(self.data[screen][algorithm][criterion][0])['lmt'][-1]) > 80:
                                self.tighten_range([screen, algorithm, criterion], 10, is_auto=True)
                        elif criterion == 'CenterLv':
                            # 100以上收紧10，否则收紧5
                            self.tighten_range([screen, algorithm, criterion],
                                               (5 if '120' in screen else 10), is_auto=True)
        return

    def read_limit_xml(self):
        """
        读取门限文件
        :return:
        """
        # 处理非UTF-8编码
        with open(self.path, 'rb') as fbobj:
            self.text = fbobj.readline()
            lsub = self.text.lower()
            if (
                b'encoding=' in lsub and
                not lsub[lsub.find(b'encoding=') + 10:].startswith(b'utf-8')
            ):
                # 有官方格式标签，类似`<?xml version="1.0" encoding="GB2312"?>`
                start = lsub.find(b'encoding=') + 10
                end = lsub.find(lsub[start - 1], start)  # self.text[start: end]为编码
                enc = self.text[start: end]
                # 重新读取
                fbobj.seek(0)
                self.text = fbobj.read()
                self.text = self.text[: start] + b'UTF-8' + self.text[end:]
                self.text = self.text.decode(enc.decode())
            else:
                # 未找到相关标签，按gbk处理
                fbobj.seek(0)
                self.text = fbobj.read()
                self.text = self.text.decode('GB2312')
            self.origin_root = Et.fromstring(self.text)
        return self

    def record_xlsx(self, path):
        """
        保存当前修改状态至指定路径
        :param path:
        :return:
        """
        wb = openpyxl.Workbook()
        sheet = wb.active
        # 样式
        common = NamedStyle(name='common')
        common.alignment = Alignment(horizontal='center', vertical='center')
        bd = Side(border_style='thin', color='000000')
        common.border = Border(bottom=bd, left=bd, right=bd, top=bd)
        common.font = Font(name='微软雅黑', size=12)
        wb.add_named_style(common)
        # 准备填充浅粉色
        pink_bkg = PatternFill('solid', fgColor='FFB6C1')
        # 列宽
        for i in 'ABCDEFG':
            col = sheet.column_dimensions[i]
            col.width = 40
        # 表头
        sheet.cell(column=1, row=1, value='测试画面').style = 'common'
        # 合并B1:C1
        sheet.cell(column=2, row=1, value='测试项').style = 'common'
        sheet.merge_cells(end_column=3, end_row=1,
                          start_column=2, start_row=1)
        sheet.cell(column=4, row=1, value='EMS').style = 'common'
        sheet.cell(column=5, row=1, value='前移').style = 'common'
        sheet.cell(column=6, row=1, value='审视').style = 'common'
        sheet.cell(column=7, row=1, value='备注').style = 'common'
        # 填入数据
        row_i = 2
        for screen, screen_item in self.data.items():
            screen_ri = row_i
            for algorithm, algorithm_item in screen_item.items():
                algorithm_ri = row_i
                for criterion, criterion_item in algorithm_item.items():
                    # 获得修改后的数据
                    updated_range = criterion_item[0]
                    for i in (2, 1):
                        if criterion_item[i]:
                            updated_range = criterion_item[i]
                            break
                    sheet.cell(column=3, row=row_i, value=criterion).style = 'common'
                    sheet.cell(column=4, row=row_i, value=criterion_item[0]).style = 'common'
                    sheet.cell(column=5, row=row_i, value=updated_range).style = 'common'
                    sheet.cell(column=6, row=row_i).style = 'common'
                    sheet.cell(column=7, row=row_i).style = 'common'
                    # 当修改与原值不同时标注颜色
                    if updated_range != criterion_item[0]:
                        sheet.cell(column=4, row=row_i).fill = pink_bkg
                        sheet.cell(column=5, row=row_i).fill = pink_bkg
                    row_i += 1
                sheet.cell(column=2, row=algorithm_ri, value=algorithm).style = 'common'
                sheet.merge_cells(end_column=2, end_row=row_i - 1,
                                  start_column=2, start_row=algorithm_ri)
            sheet.cell(column=1, row=screen_ri, value=screen).style = 'common'
            sheet.merge_cells(end_column=1, end_row=row_i - 1,
                              start_column=1, start_row=screen_ri)
        # 写入文档
        wb.save(path)
        return self

    def tighten_range(self, path, size, is_auto=False):
        """
        根据键路径与大小去收紧相应门限
        :param path:
        :param size:
        :param is_auto:
        :return:
        """
        # 获得当前门限
        target_range = self.data
        for k in path:
            target_range = target_range[k]
        # 解析门限
        limit_dict = self.parse_range(target_range[0])
        # 仅当门限成对时处理
        if len(limit_dict['lmt']) != 2:
            return self
        # 确认收紧后门限精度
        dightn = 0
        # 所有门限内的最大精度
        for n in limit_dict['lmt']:
            if '.' in n and len(n) - n.index('.') > dightn + 1:
                dightn = len(n) - n.index('.') - 1
        # 收紧门限的最大精度
        rare_dights = abs(size) * 10**dightn
        rare_dights = round(rare_dights - int(rare_dights), self.MAX_DIGHT_N)
        while rare_dights:
            dightn += 1
            rare_dights *= 10
            rare_dights = round(rare_dights - int(rare_dights), self.MAX_DIGHT_N)
        # 最大精度极值
        if dightn > self.MAX_DIGHT_N:
            dightn = self.MAX_DIGHT_N
        # 计算
        num1 = float(limit_dict['lmt'][0])
        num2 = float(limit_dict['lmt'][1])
        direction = -1 if num1 > num2 else 1
        # 通常以一个很大的正数或很小的负数来代表无穷，此时不对它进行收紧
        if -self.INF_LIMIT < num1 < self.INF_LIMIT:
            num1 = round(num1 + direction * size, dightn)
        if -self.INF_LIMIT < num2 < self.INF_LIMIT:
            num2 = round(num2 - direction * size, dightn)
        # 防呆：收紧后大小关系发生变化则停止操作
        if (num1 - num2) * direction > 0:
            return self
        # 组装新的门限
        target_range[1 if is_auto else 2] = f'{limit_dict["lp"]}{num1:g}{limit_dict["cp"]}{num2:g}{limit_dict["rp"]}'
        return self

    def write_updated_xml(self, path):
        """
        将更改后的xml写到指定路径
        :param path:
        :return:
        """
        updated_root = Et.fromstring(self.text)
        for screen_item in updated_root.findall('.//TestItem[@Name]'):
            screen = screen_item.get('Name')
            if screen in self.data:
                screen_inner = self.data[screen]
                for algorithm_item in screen_item.findall('.//AlgorithmItem[@Name]'):
                    algorithm = algorithm_item.get('Name')
                    if algorithm in screen_inner:
                        algorithm_inner = screen_inner[algorithm]
                        for criterion_item in algorithm_item.findall('.//CriterionItem[@Name]'):
                            criterion = criterion_item.get('Name')
                            if criterion in algorithm_inner:
                                item_range = algorithm_inner[criterion]
                                # 存在修改
                                if item_range[2] or item_range[1]:
                                    updated_range = item_range[2] if item_range[2] else item_range[1]
                                    # 修改后与修改前有差别
                                    if updated_range != item_range[0]:
                                        if 'Range' in criterion_item.keys():
                                            criterion_item.set('Range', updated_range)
                                        else:
                                            criterion_item.text = updated_range
        Et.ElementTree(updated_root).write(path, encoding='GB2312')
        return self
