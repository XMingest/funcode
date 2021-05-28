# -*- coding: utf-8 -*-
import sys
import xml.etree.ElementTree as Et
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, NamedStyle, Side


def get_config_item(path):
    """
    存在结构如下的xml
    ```xml
<?xml version="1.0" encoding="GB2312"?>
<MachineTestStandard>
<MachineType>MediaTest 9508B</MachineType>
<ReleaseVersion>1.021</ReleaseVersion>
<ReleaseDate>2020/08/07</ReleaseDate>
<TestItems>
    <TestItem Name="White90Hz" ChromaType="Rvs">
        <AlgorithmItem Name="Chroma" Desc="Chroma">
            <TestStandards Desc="多媒体V6.0" ProductType="Noah" IsActive="1">
                <CriterionItem Name="CCT" ReverseName1="CCT" ReverseName2="CCTTolerance"
                               Desc="色温中心值" Remark="" Type="2" Range="[6000:8000]" />
                <CriterionItem Name="Cx" ReverseName1="C1" ReverseName2="C1Tolerance"
                               Desc="色坐标x" Remark="" Type="2" Range="[0.285:0.325]" />
                <CriterionItem Name="Cy" ReverseName1="C2" ReverseName2="C2Tolerance"
                               Desc="色坐标y" Remark="" Type="2" Range="[0.301:0.341]" />
            </TestStandards>
        </AlgorithmItem>
    </TestItem>
</TestItems>
</MachineTestStandard>

    ```
    则从中得到
    ```python
    {'White90Hz': {'Chroma': {
                                'CCT': '[6000:8000]',
                                'Cx': '[0.285:0.325]',
                                'Cy': '[0.301:0.341]',
    }}}
    ```
    :param path:
    :return:
    """
    try:
        root = Et.parse(path).getroot()
    except ValueError:
        # 处理非UTF-8编码
        with open(path, 'rb') as fbobj:
            text = fbobj.readline()
            lsub = text.lower()
            if (
                b'encoding=' in lsub and
                not lsub[lsub.find(b'encoding=') + 10:].startswith(b'utf-8')
            ):
                # 有官方格式标签，类似`<?xml version="1.0" encoding="GB2312"?>`
                start = lsub.find(b'encoding=') + 10
                end = lsub.find(lsub[start - 1], start)  # text[start: end]为编码
                enc = text[start: end]
                # 重新读取
                fbobj.seek(0)
                text = fbobj.read()
                text = text[: start] + b'UTF-8' + text[end:]
                root = Et.fromstring(text.decode(enc.decode()))
            else:
                # 未找到相关标签，按gbk处理
                fbobj.seek(0)
                text = fbobj.read()
                root = Et.fromstring(text.decode('GBK'))
    # 格式化数据
    config_data = {}
    for test_item in root.findall('.//TestItem'):
        test_screen = test_item.attrib['Name']
        for algorithm_item in test_item.findall('.//AlgorithmItem'):
            algorithm = algorithm_item.attrib['Name']
            for criterion_item in algorithm_item.findall('.//CriterionItem'):
                criterion = criterion_item.attrib['Name']
                if 'Range' in criterion_item.attrib:
                    limit = criterion_item.attrib['Range']
                else:
                    limit = criterion_item.text
                config_data.setdefault(test_screen, {}) \
                    .setdefault(algorithm, {})[criterion] = limit
    return config_data


def limit_xlsx_gen(data, path):
    """
    通过数据在指定路径生成excel表格文件
    :param data:
    :param path:
    :return:
    """
    wb = Workbook()
    sheet = wb.active
    # 样式
    common = NamedStyle(name='common')
    common.alignment = Alignment(horizontal='center', vertical='center')
    bd = Side(border_style='thin', color='000000')
    common.border = Border(bottom=bd, left=bd, right=bd, top=bd)
    common.font = Font(name='微软雅黑', size=12)
    wb.add_named_style(common)
    # 列宽
    for i in 'ABCDE':
        col = sheet.column_dimensions[i]
        col.width = 40
    # 表头
    sheet.cell(column=1, row=1, value='测试画面').style = 'common'
    # 合并B1:C1
    sheet.cell(column=2, row=1, value='测试项').style = 'common'
    sheet.merge_cells(end_column=3, end_row=1,
                      start_column=2, start_row=1)
    sheet.cell(column=4, row=1, value='门限(项目名称：)').style = 'common'
    sheet.cell(column=5, row=1, value='备注：').style = 'common'
    # 填入数据
    row_i = 2
    for test_screen in data:
        screen_ri = row_i
        for algorithm in data[test_screen]:
            algorithm_ri = row_i
            for criterion in data[test_screen][algorithm]:
                sheet.cell(column=3, row=row_i, value=criterion).style = 'common'
                sheet.cell(column=4, row=row_i,
                           value=data[test_screen][algorithm][criterion]).style = 'common'
                row_i += 1
            sheet.cell(column=2, row=algorithm_ri, value=algorithm).style = 'common'
            sheet.merge_cells(end_column=2, end_row=row_i - 1,
                              start_column=2, start_row=algorithm_ri)
        sheet.cell(column=1, row=screen_ri, value=test_screen).style = 'common'
        sheet.merge_cells(end_column=1, end_row=row_i - 1,
                          start_column=1, start_row=screen_ri)
    # 写入文档
    wb.save(path)
    return


if __name__ == '__main__':
    try:
        limit_path = Path(sys.argv[1])
        if limit_path.is_file() and limit_path.suffix == '.xml':
            limit_xlsx_gen(get_config_item(limit_path), limit_path.with_suffix('.xlsx'))
    except IndexError:
        import tkinter
        from tkinter import filedialog, messagebox

        tkinter.Tk().withdraw()
        limit_path = Path(filedialog.askopenfilename(filetypes=[('门限文件', '*.xml')], title='请选择门限文件'))
        if not limit_path.is_file() or limit_path.suffix != '.xml':
            messagebox.showinfo('提示', '未选择合法XML文件')
            sys.exit()
        save_path = Path(filedialog.asksaveasfilename(defaultextension='.xlsx',
                                                      filetypes=[('表格', '*.xlsx')],
                                                      title='请选择保存文件'))
        if save_path.suffix != '.xlsx':
            messagebox.showinfo('提示', '未指定合法保存路径')
            sys.exit()
        limit_xlsx_gen(get_config_item(limit_path), save_path)
        messagebox.showinfo('提示', '导出成功！')
