# -*- coding: utf-8 -*-
import json
from pathlib import Path

import openpyxl


class MonthRecord:
    def __init__(self):
        # config
        cfg_fp = Path('config.json')
        if not cfg_fp.is_file():
            with open(cfg_fp, 'w', encoding='UTF-8') as cfg_fobj:
                json.dump({'names': []}, cfg_fobj)
        with open(cfg_fp, 'r', encoding='UTF-8') as cfg_fobj:
            self.config = json.load(cfg_fobj)
        # data
        data_fp = Path('data.json')
        if not data_fp.is_file():
            with open(data_fp, 'w', encoding='UTF-8') as data_fobj:
                json.dump({}, data_fobj)
        with open(data_fp, 'r', encoding='UTF-8') as data_fobj:
            self.data = json.load(data_fobj)
        # init
        self.title = {}
        return

    def add_name(self, name):
        """
        添加名字
        :param name:
        :return:
        """
        arr = set(self.config['names'])
        arr.add(name)
        self.config['names'] = list(arr)
        return self

    def add_names(self, names):
        """
        添加名字
        :param names:
        :return:
        """
        arr = set(self.config['names'])
        arr |= set(names)
        self.config['names'] = list(arr)
        return self

    def append_data(self, workbook):
        """
        添加工作薄数据
        :param workbook:
        :return:
        """
        wb = openpyxl.load_workbook(workbook)
        ws = wb.active
        name_col = None
        title = None
        for row in ws.rows:
            if name_col and row[name_col].value in self.config['names']:
                self.data[title][row[name_col].value] = {f'{self.title[t]}': f'{row[t].value}' for t in self.title}
            elif name_col is None:
                for cell in row:
                    if cell.value == '姓名':
                        name_col = cell.column - 1
                        title = Path(workbook).stem
                        self.data[title] = {}
                        for row_item in ws[cell.row + 1]:
                            self.title[row_item.column - 1] = row_item.value
                        for row_item in ws[cell.row]:
                            if self.title[row_item.column - 1] is None:
                                self.title[row_item.column - 1] = row_item.value
                        break
        wb.close()
        return self

    def update(self):
        """
        更新数据
        :return:
        """
        with open('config.json', 'w',
                  encoding='UTF-8') as cfg_fobj, open('data.json', 'w',
                                                      encoding='UTF-8') as data_fobj:
            json.dump(self.config, cfg_fobj, ensure_ascii=False)
            json.dump(self.data, data_fobj, ensure_ascii=False)
        return self
